"""

.. Hint::
    openpyxl 封装的excel操作

    需要安装 openpyxl

    pip install openpyxl


.. literalinclude:: ../../../example/excel_openpyxl_使用例子.py
    :language: python
    :caption: 代码示例
    :linenos:

"""

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块
from openpyxl.styles import Border, Side
from openpyxl.drawing.image import Image



class Excel():
    def 打开Excel(self, 文件路径):
        pass
        self.wb = load_workbook(filename=文件路径)
        self.st = self.wb.active

    def 创建空白工作簿(self):
        pass
        self.wb = Workbook()
        self.st = self.wb.active

    def 置当前sheet名称(self, 名称):
        self.st.title = 名称

    def 取当前sheet对象(self):
        return self.st

    def 取当前sheet名称(self):
        return self.st.title

    def 创建Sheet(self, 名称):
        pass
        self.st = self.wb.create_sheet(名称)

    def 取所有sheet名称(self):
        return self.wb.sheetnames

    def 保存(self, 文件路径):
        self.wb.save(文件路径)

    def 置当前sheet(self, 名称):
        pass
        self.st = self.wb[名称]

    def 取某列的所有内容(self, 第几列):
        rows = self.st.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.st.cell(row=i, column=第几列).value
            column_data.append(cell_value)
        return column_data

    def 取某行的所有内容(self, 第几行):
        columns = self.st.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.st.cell(row=第几行, column=i).value
            row_data.append(cell_value)
        return row_data

    def 置图片(self, 单元格位置="A1", 宽=100, 高=100, 图片路径=""):
        img = Image(图片路径)
        img.width = 宽
        img.height = 高
        self.st.add_image(img, 单元格位置)

    def 置内容(self, 行, 列, 内容):
        pass
        self.st.cell(row=行, column=列, value=内容)

    def 取内容(self, 行, 列):
        pass
        return self.st.cell(row=行, column=列).value

    def 取行数(self):
        pass
        return self.st.max_row

    def 取列数(self):
        pass
        return self.st.max_column

    def 置列宽(self, 列, 宽度):
        "列 A B C D E "
        self.st.column_dimensions[列].width = 宽度

    def 置行高(self, 行, 高度):
        "行 1 2 3 4 5 "
        self.st.row_dimensions[行].height = 高度

    def 合并单元格(self, 范围="A1:A10"):
        "A2:A10 合并A2到A10之间的单元格"
        self.st.merge_cells(范围)

    def 置单元格背景颜色(self, 行: int, 列: int, 颜色="ffffff"):
        '# 1开始 注意：fill_type为填充类型，如果不写的话，则没有效果。 默认纯色填充 颜色代号 白=ffffff 红=ff0000 黄=ffff00 绿=33ff00 蓝=0000ff 黑=000000 紫=9900ff'
        fill = PatternFill(fill_type='solid', fgColor=颜色)
        self.st.cell(row=行, column=列).fill = fill

    def 置单元格对齐方式(self, 行: int, 列: int, 对齐方式=Alignment(horizontal="justify", vertical="justify")):
        """
        horizontal代表水平方向，可以左对齐left，还有居中center和右对齐right，分散对齐distributed，跨列居中centerContinuous，两端对齐justify，填充fill，常规general

        vertical代表垂直方向，可以居中center，还可以靠上top，靠下bottom，两端对齐justify，分散对齐distributed

        另外还有自动换行：wrap_text，这是个布尔类型的参数，这个参数还可以写作wrapText


        """
        self.st.cell(row=行, column=列).alignment = 对齐方式

    def 置单元格字体风格(self, 行: int, 列: int,
                 字体名称=u"微软雅黑",
                 字体大小=11,
                 加粗=False,
                 斜体=False,
                 删除线=False,
                 颜色="ffffff"):
        ' 白=ffffff 红=ff0000 黄=ffff00 绿=33ff00 蓝=0000ff 黑=000000 紫=9900ff'
        font = Font(字体名称, size=字体大小, bold=加粗, italic=斜体, strike=删除线, color=颜色)
        self.st.cell(row=行, column=列).font = font

    def 置单元格边框(self, 行: int, 列: int,
               left=Side(border_style="thin", color="000000"),
               right=Side(border_style="thin", color="000000"),
               top=Side(border_style="thin", color="000000"),
               bottom=Side(border_style="thin", color="000000"),
               ):
        # dashDot,
        # dashDotDot,
        # dashed,
        # dotted,
        # double,
        # hair,
        # medium,
        # mediumDashDot,
        # mediumDashDotDot,
        # mediumDashed,
        # slantDashDot,
        # thick,
        # thin,
        border = Border(left=left,
                        right=right,
                        top=top,
                        bottom=bottom)
        self.st.cell(row=行, column=列).border = border

    def 删除当前sheet(self, sheet名称=None):
        '不设置名称则删除当前sheet 若设置则删除指定的sheet'
        try:
            if sheet名称 == None:
                self.wb.remove(self.st)
            else:
                del self.wb[sheet名称]
        except IndexError:
            print('最少保留一个表格对象')

    def 关闭(self):
        self.wb.close()

