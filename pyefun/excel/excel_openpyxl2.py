#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2020-11-21 21:12
# @Author : Nuonuo
# @Site : 
# @File : 表格.py
# @Software: PyCharm


import openpyxl  # 表格


# 表格openpyxl================================================================
class 表格:
    def __init__(self):
        print('初始化 openpyxl 成功')
        print('行数0开始   列数1开始')

    # '''
    # 在使用openpyxl模块前，需要了解openpyx中名称的概念：
    # 在openpyxl中，主要用到三个概念：Workbooks，Sheets，Cells
    # Workbook：指一个excel工作表；
    # Sheet：指工作表中的一张表；
    # Cell：指表中的一个单元格。
    # openpyx在读或写时顺序为：打开Workbook，定位Sheet，操作Cell
    # '''
    def 创建_工作簿对象(self):
        '''如果写入中文为乱码，可添加参数encoding = '''
        self.__wb = openpyxl.Workbook()
        print('初始化workbok成功')
        print(self.__wb)

    def 创建_工作表(self, 工作表名称, 位置=None):
        '默认在最后添加一个工作表 也可以选择位置插入  openpyxl模块暂不支持xls格式文件'
        self.__sheet = self.__wb.create_sheet(工作表名称, 位置)

    def 读取_工作簿对象(self, path, 只读=False):
        if path[-4:] == 'xlsx':
            self.__wb = openpyxl.load_workbook(path, read_only=只读)
            print('初始化workbok成功')
            print(self.__wb)
            self.设置_默认工作表对象()
        else:
            print(path, '只支持xlsx格式文件')

    def 设置_默认工作表对象(self):
        '默认获取第一个表格对象'
        self.__sheet = self.__wb[self.__wb.sheetnames[0]]
        print('初始化worksheet成功')
        print(self.__sheet)

    def 查看_所有工作表名字(self):
        return self.__wb.sheetnames

    def 选择表(self, 表名):
        ''' #sheet1 = book.worksheets[n]、sheet2 = book['工作表名称']、sheet3 = book[book.sheetnames[n]]'''
        print('初始化worksheet成功')
        self.__sheet = self.__wb[表名]

    def 获取_当前工作表_名称(self):
        return self.__sheet.title

    def 获取_关键字_行内容(self, 关键字, 关键列, 只取id=False):
        '遍历查询调用内部行数'
        最大行数 = self.__sheet.max_row
        data = None
        for i in range(1, 最大行数 + 1):
            cell_value = self.__sheet.cell(row=i, column=关键列).value
            if cell_value == 关键字:
                if 只取id:
                    data = i
                else:
                    data = self.获取_某行所有值(i)
                break
            # data.append(cell_value)
        if data == None:
            return '没有查询数据'
        return data

    def 获取_关键字_列内容(self, 关键字, 关键列, 只取id=False):
        '遍历查询调用内部行数'
        最大列数 = self.__sheet.max_column
        data = None
        for i in range(1, 最大列数 + 1):
            cell_value = self.__sheet.cell(row=关键列, column=i).value
            if cell_value == 关键字:
                if 只取id:
                    data = i
                else:
                    data = self.获取_某列的所有值(i)
                break
            # data.append(cell_value)
        if data == None:
            return '没有查询数据'
        return data

    def 获取_某列的所有值(self, 列):
        rows = self.__sheet.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.__sheet.cell(row=i, column=列).value
            column_data.append(cell_value)
        return column_data

    def 获取_某行所有值(self, 行):
        columns = self.__sheet.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.__sheet.cell(row=行, column=i).value
            row_data.append(cell_value)
        return row_data

    def 重命名_当前工作表(self, 新表名):
        self.__sheet.title = 新表名

    def 获取_工作表行数(self):
        return self.__sheet.max_row

    def 获取_工作表列数(self):
        return self.__sheet.max_column

    def 获取_单元格内容(self, 行: int, 列: int):
        '''#获取某一单元格内容：cell = sheet.cell(1,2).value、sheet['单元格'].value例如sheet['B1'].value
print(sheet.cell(20,20).value)'''
        return self.__sheet.cell(行, 列).value

    def 写入_单元格内容(self, 行: int, 列: int, 值):
        '行和列从1开始'
        self.__sheet.cell(row=行, column=列, value=值)

    def 获取_当前表格所有内容(self):
        # 按行读取 工作表的内容
        for row in self.__sheet.rows:
            for cell in row:
                print(cell.value, end=' ')
            print('')

    def 填充_单元格颜色(self, 行: int, 列: int, 颜色):
        '# 1开始 注意：fill_type为填充类型，如果不写的话，则没有效果。 默认纯色填充 颜色代号 白=ffffff 红=ff0000 黄=ffff00 绿=33ff00 蓝=0000ff 黑=000000 紫=9900ff'
        if 颜色 == '白':
            ls = 'ffffff'
        elif 颜色 == '黑':
            ls = '000000'
        elif 颜色 == '红':
            ls = 'ff0000'
        elif 颜色 == '黄':
            ls = 'ffff00'
        elif 颜色 == '绿':
            ls = '33ff00'
        elif 颜色 == '蓝':
            ls = '0000ff'
        else:
            ls = 'ffffff'
            print('输入颜色错误 恢复默认-白色')
        fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=ls)
        self.__sheet.cell(row=行, column=列).fill = fill

    def 保存_工作簿(self, 文件名):
        self.__wb.save(文件名)

    def 删除_当前工作表(self, 工作表名=None):
        '两种方法删除 一种是默认删除当前 一种可以选择'
        try:
            if 工作表名 == None:
                self.__wb.remove(self.__sheet)
            else:
                del self.__wb[工作表名]
        except IndexError:
            print('最少保留一个表格对象')

    def 退出(self):
        self.__wb.close()
